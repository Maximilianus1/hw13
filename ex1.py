import openpyxl
wb_all=openpyxl.Workbook()
wb1=openpyxl.load_workbook('1111.xlsx')
wb2=openpyxl.load_workbook('2222.xlsx')
wb3=openpyxl.load_workbook('3333.xlsx')
ws_all=wb_all.active
ws1=wb1.active
ws2=wb2.active
ws3=wb3.active
lst1=[]
lst2=[]
lst3=[]
lst_all=[]
for row in ws1.values:
    for value in row:
        lst1.append(value)
for row in ws2.values:
    for value in row:
        lst2.append(value)
for row in ws3.values:
    for value in row:
        lst3.append(value)
lst1=sorted(lst1)
lst2=sorted(lst2)
lst3=sorted(lst3)
lst_all=[lst1,lst2,lst3]
lst_sort=[]
for i in range(len(lst_all)):
    for j in range(len(lst_all[i])):
        lst_sort.append(lst_all[i][j])
lst_sort=sorted(lst_sort)
lst_sort=list(reversed(lst_sort))
o=0
for i in range(len(lst_all)):
    for j in range(len(lst_all[i])):
        lst_all[i][j]=lst_sort[o]
        ws_all.cell(row=i+1,column=j+1, value=lst_sort[o]).font=openpyxl.styles.Font(name='Comic Sans MS', charset=204, family=2.0, sz=16.0)
        ws_all.cell(row=i+1,column=j+1, value=lst_sort[o]).border=openpyxl.styles.Border(top=openpyxl.styles.borders.Side(border_style="thin"),left=openpyxl.styles.borders.Side(border_style="thin"),right=openpyxl.styles.borders.Side(border_style="thin"),bottom=openpyxl.styles.borders.Side(border_style="thin"))
        o+=1
wb_all.save('finish.xlsx')
wb1.close()
wb2.close()
wb3.close()
wb_all.close()
